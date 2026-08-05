import os
import io
import json
import time
import html
import logging
import threading
from datetime import datetime
from zoneinfo import ZoneInfo
from http.server import BaseHTTPRequestHandler, HTTPServer
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, InputFile, Update
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes

# Configuración de logs
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Configuración de variables de entorno
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
GITHUB_REPO = os.getenv("GITHUB_REPO", "tu_usuario/bot-dakhla-atlantique")
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
JSON_RAW_URL = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/registro.json"
GROQ_MODEL = "llama-3.3-70b-versatile"
GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"

ZONA_CANARIAS = ZoneInfo("Atlantic/Canary")
HORA_ENVIO_OBJETIVO = 8  # 8:00 hora de Canarias

FLAG_MAP = {
    "arabe": "🇲🇦", "árabe": "🇲🇦",
    "frances": "🇫🇷", "francés": "🇫🇷",
    "espanol": "🇪🇸", "español": "🇪🇸",
    "ingles": "🇬🇧", "inglés": "🇬🇧",
    "vídeo": "📺", "video": "📺",
}

MESES_NOMBRE = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
}


def bandera_para(etiqueta: str) -> str:
    """Devuelve el emoji más adecuado para una etiqueta (idioma, fuente de radio o podcast)."""
    clave = (etiqueta or "").strip().lower()
    if clave in FLAG_MAP:
        return FLAG_MAP[clave]
    if clave.startswith("radio"):
        return "📻"
    if clave:
        return "🎙️"  # fuentes de podcast (nombre del programa)
    return "🌐"


def obtener_datos_registro() -> dict:
    """
    Descarga registro.json — la fuente de datos estructurada que escribe main.py —
    y la convierte a la forma {año: {mes: [items]}} que usa el resto del bot.
    Al ser JSON real (no texto a analizar con expresiones regulares), esto no se
    rompe aunque cambiemos el diseño visual del reporte de Telegram en el futuro.
    """
    try:
        response = requests.get(JSON_RAW_URL, timeout=10)
        if response.status_code != 200:
            logger.error(f"Error al obtener registro.json: HTTP status {response.status_code}")
            return {}
        registro = response.json()
    except Exception as e:
        logger.error(f"Excepción al obtener/leer registro.json: {e}")
        return {}

    datos = {}
    for fecha, contenido_dia in registro.items():
        partes = fecha.split("-")
        if len(partes) != 3:
            continue
        anio, mes, _ = partes
        datos.setdefault(anio, {}).setdefault(mes, [])

        for item in contenido_dia.get("items", []):
            etiqueta = item.get("idioma") or item.get("etiqueta") or item.get("categoria") or "Otros"
            datos[anio][mes].append({
                "fecha": fecha,
                "bandera": bandera_para(etiqueta),
                "etiqueta": etiqueta,
                "titular": item.get("titular", ""),
                "link": item.get("link", "#")
            })

    # Orden más reciente primero dentro de cada mes, igual que antes
    for anio in datos:
        for mes in datos[anio]:
            datos[anio][mes].sort(key=lambda it: it["fecha"], reverse=True)

    return datos


TEXTO_AYUDA = (
    "🤖 *Bot Dakhla Atlantique — Ayuda e Información*\n\n"
    "📌 *Registro Histórico*\n"
    "Consulta el archivo histórico de noticias, navegando año a año y mes a mes.\n\n"
    "📤 *Exportar a Excel*\n"
    "Descarga en un archivo Excel (.xlsx) las noticias registradas de cualquier mes.\n\n"
    "⚓ *Resumen Mensual*\n"
    "Genera, cuando tú lo pidas, un resumen del mes completo en dos versiones: "
    "una puramente factual y otra con perspectiva de analista (sin valoraciones). "
    "No se envía nunca automáticamente.\n\n"
    "🔄 *Forzar Reporte*\n"
    "Dispara manualmente la ejecución del bot en GitHub Actions al momento.\n\n"
    "📊 *Estado del Sistema*\n"
    "Comprueba el estado de la última ejecución y salud general del bot.\n\n"
    "El reporte diario se envía automáticamente todos los días a las 08:00 hora de Canarias."
)


async def comando_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra el panel de control interactivo con todas las opciones."""
    keyboard = [
        [InlineKeyboardButton("🔄 Forzar Reporte", callback_data="menu_forzar")],
        [InlineKeyboardButton("📊 Estado del Sistema", callback_data="menu_estado")],
        [InlineKeyboardButton("📜 /registro-historico", callback_data="menu_historico")],
        [InlineKeyboardButton("📊 /exportar", callback_data="menu_exportar")],
        [InlineKeyboardButton("⚓ /resume_mes", callback_data="menu_resumenmes")],
        [InlineKeyboardButton("❓ /ayuda", callback_data="menu_ayuda")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    texto = "🎛️ *Panel de Control — Dakhla Atlantique*\n\nSelecciona una opción del menú:"

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.edit_text(texto, reply_markup=reply_markup, parse_mode="Markdown")
    elif update.message:
        await update.message.reply_text(texto, reply_markup=reply_markup, parse_mode="Markdown")


async def comando_ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton("« Volver al Menú Principal", callback_data="menu_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.edit_text(TEXTO_AYUDA, reply_markup=reply_markup, parse_mode="Markdown")
    elif update.message:
        await update.message.reply_text(TEXTO_AYUDA, reply_markup=reply_markup, parse_mode="Markdown")


def preguntar_ia(prompt: str):
    """Llama a Groq (mismo proveedor que usa main.py para el resumen diario)."""
    if not GROQ_API_KEY:
        return None
    try:
        res = requests.post(
            GROQ_URL,
            headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
            json={
                "model": GROQ_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.5,
            },
            timeout=60
        )
        if res.status_code != 200:
            logger.error(f"Error llamando a la IA (Groq): HTTP {res.status_code} - {res.text}")
            return None
        return res.json()["choices"][0]["message"]["content"]
    except Exception as e:
        logger.error(f"Error llamando a la IA (Groq): {e}")
        return None


def construir_texto_titulares_mes(items_mes):
    """Construye el listado de titulares del mes, en orden cronológico, para pasárselo a la IA."""
    items_ordenados = sorted(items_mes, key=lambda it: it["fecha"])
    return "\n".join(f"- [{it['fecha']}] {it['titular']}" for it in items_ordenados)


def generar_resumen_mensual_general(texto_titulares, mes_nombre, anio):
    """Resumen puramente extractivo del mes: sin interpretación ni añadidos."""
    prompt = f"""
    A continuación tienes TODOS los titulares recopilados durante {mes_nombre} de {anio}
    sobre el Puerto de Dakhla Atlantique, en orden cronológico.

    Redacta un resumen factual (varios párrafos si hace falta) que sintetice
    ÚNICAMENTE la información contenida en estos titulares.

    Reglas estrictas:
    - No añadas opiniones, valoraciones, interpretaciones ni conclusiones propias.
    - No completes con datos que no estén en la lista.
    - Limítate a describir de forma neutra qué se ha publicado durante el mes.
    - Responde solo con texto plano, sin Markdown ni HTML.

    Titulares del mes:
    {texto_titulares}
    """
    return preguntar_ia(prompt)


def generar_resumen_mensual_analista(texto_titulares, mes_nombre, anio):
    """Resumen del mes con perspectiva de analista de inteligencia: conecta
    hechos y detecta patrones, pero sin emitir juicios ni valoraciones."""
    prompt = f"""
    A continuación tienes TODOS los titulares recopilados durante {mes_nombre} de {anio}
    sobre el Puerto de Dakhla Atlantique, en orden cronológico.

    Redacta un resumen con la perspectiva de un analista de inteligencia: organiza
    la información por temas o actores relevantes, conecta hechos relacionados
    entre sí y señala patrones o líneas de desarrollo que se repitan a lo largo
    del mes.

    Reglas estrictas:
    - NO emitas valoraciones, juicios personales, opiniones ni predicciones.
    - No califiques los hechos como positivos, negativos, preocupantes, etc.
    - Limítate a mostrar relaciones y patrones objetivos entre los hechos
      publicados, sin añadir información que no esté en los titulares.
    - Responde solo con texto plano, sin Markdown ni HTML.

    Titulares del mes:
    {texto_titulares}
    """
    return preguntar_ia(prompt)


async def comando_resumen_mensual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Punto de entrada: muestra los años disponibles para el resumen mensual."""
    datos = obtener_datos_registro()
    if not datos:
        msg = "⚠️ No se pudo acceder al registro histórico en este momento. Inténtalo más tarde."
        if update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.message.reply_text(msg)
        return

    keyboard = []
    for year in sorted(datos.keys(), reverse=True):
        keyboard.append([InlineKeyboardButton(f"📂 Año {year}", callback_data=f"resmes_year_{year}")])

    keyboard.append([InlineKeyboardButton("« Volver al Menú Principal", callback_data="menu_main")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    texto = "⚓ *Resúmenes Mensuales — Puerto de Dakhla Atlantique*\n\nSelecciona el año:"

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.edit_text(texto, reply_markup=reply_markup, parse_mode="Markdown")
    elif update.message:
        await update.message.reply_text(texto, reply_markup=reply_markup, parse_mode="Markdown")
    datos = obtener_datos_registro()
    if not datos:
        msg = "⚠️ No se pudo acceder al registro histórico en este momento. Inténtalo más tarde."
        if update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.message.reply_text(msg)
        return

    keyboard = []
    for year in sorted(datos.keys(), reverse=True):
        keyboard.append([InlineKeyboardButton(f"📂 Año {year}", callback_data=f"year_{year}")])
    
    keyboard.append([InlineKeyboardButton("« Volver al Menú Principal", callback_data="menu_main")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    texto = "📌 *Registro Histórico Dakhla Atlantique*\n\nSelecciona un año para consultar los meses disponibles:"

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.edit_text(texto, reply_markup=reply_markup, parse_mode="Markdown")
    elif update.message:
        await update.message.reply_text(texto, reply_markup=reply_markup, parse_mode="Markdown")


async def comando_exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    datos = obtener_datos_registro()
    if not datos:
        msg = "⚠️ No se pudo acceder al registro histórico en este momento. Inténtalo más tarde."
        if update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.message.reply_text(msg)
        return

    keyboard = []
    for year in sorted(datos.keys(), reverse=True):
        keyboard.append([InlineKeyboardButton(f"📂 Año {year}", callback_data=f"exp_year_{year}")])
    
    keyboard.append([InlineKeyboardButton("« Volver al Menú Principal", callback_data="menu_main")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    texto = "📤 *Exportar registro a Excel*\n\nSelecciona el año:"

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.edit_text(texto, reply_markup=reply_markup, parse_mode="Markdown")
    elif update.message:
        await update.message.reply_text(texto, reply_markup=reply_markup, parse_mode="Markdown")


def disparar_workflow_github(modo="manual"):
    """
    Llama a la API de GitHub para lanzar daily_bot.yml (workflow_dispatch).
    modo="manual" -> el usuario pulsó "Forzar Reporte": ignora la hora y el
                      filtro de "ya enviado hoy" (siempre foto actual).
    modo="automatico" -> lo dispara el reloj interno de este mismo bot a las
                      8:00 de Canarias: se comporta como el envío programado
                      normal (respeta hora y evita duplicados).
    Devuelve (ok: bool, mensaje: str).
    """
    if not GITHUB_TOKEN:
        return False, "⚠️ Error: No se ha configurado GITHUB_TOKEN en las variables de entorno."

    url = f"https://api.github.com/repos/{GITHUB_REPO}/actions/workflows/daily_bot.yml/dispatches"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }
    try:
        response = requests.post(
            url,
            json={"ref": "main", "inputs": {"modo": modo}},
            headers=headers,
            timeout=10
        )
        if response.status_code == 204:
            return True, "🤖 Estoy trabajando en ello, en unos minutos te facilito la información. Siéntate y tómate un café. ☕"
        return False, f"⚠️ No se pudo disparar el workflow (Código HTTP {response.status_code})."
    except Exception as e:
        return False, f"❌ Excepción al conectar con GitHub: {e}"


def iniciar_reloj_disparo_diario():
    """
    Hilo de fondo (Render mantiene este proceso vivo 24h con ayuda de
    UptimeRobot): cada minuto comprueba si son las 8:00 en Canarias y, de ser
    así, dispara el reporte diario mediante la API de GitHub, sin depender
    del 'schedule' de GitHub Actions (que hemos comprobado que puede
    retrasarse horas o directamente no dispararse en cuentas gratuitas).
    """
    ultimo_dia_disparado = None
    while True:
        try:
            ahora = datetime.now(ZONA_CANARIAS)
            hoy = ahora.strftime("%Y-%m-%d")
            if ahora.hour == HORA_ENVIO_OBJETIVO and hoy != ultimo_dia_disparado:
                logger.info(f"Reloj interno: son las {ahora.strftime('%H:%M')} en Canarias, disparando el reporte diario.")
                ok, _ = disparar_workflow_github(modo="automatico")
                if ok:
                    ultimo_dia_disparado = hoy
                else:
                    logger.error("El reloj interno no pudo disparar el workflow; se reintentará en el siguiente minuto.")
        except Exception as e:
            logger.error(f"Error en el reloj de disparo diario: {e}")
        time.sleep(60)


async def comando_forzar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Dispara manualmente el workflow de GitHub Actions (botón/comando del usuario)"""
    query = update.callback_query if update.callback_query else None
    if query:
        await query.answer("Lanzando reporte...")

    _, msg = disparar_workflow_github(modo="manual")

    keyboard = [[InlineKeyboardButton("« Volver al Menú Principal", callback_data="menu_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if query:
        await query.message.edit_text(msg, reply_markup=reply_markup, parse_mode="Markdown")
    else:
        await update.message.reply_text(msg, reply_markup=reply_markup, parse_mode="Markdown")


async def comando_estado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Consulta el estado de la última ejecución en GitHub Actions"""
    query = update.callback_query if update.callback_query else None
    if query:
        await query.answer("Comprobando estado...")

    headers = {"Accept": "application/vnd.github.v3+json"}
    if GITHUB_TOKEN:
        headers["Authorization"] = f"token {GITHUB_TOKEN}"

    url = f"https://api.github.com/repos/{GITHUB_REPO}/actions/runs?per_page=1"
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            runs = response.json().get("workflow_runs", [])
            if runs:
                run = runs[0]
                estado = run.get("conclusion") or run.get("status")
                fecha = run.get("created_at", "")[:19].replace("T", " ")
                html_url = run.get("html_url")
                
                icono_estado = "✅" if estado == "success" else ("⏳" if estado in ["queued", "in_progress"] else "❌")
                msg = (
                    f"📊 *Estado del Sistema — GitHub Actions*\n\n"
                    f"{icono_estado} **Último estado:** `{estado}`\n"
                    f"🗓️ **Fecha:** `{fecha} UTC`\n\n"
                    f"🔗 [Ver detalles en GitHub]({html_url})"
                )
            else:
                msg = "ℹ️ No se encontraron ejecuciones recientes en el repositorio."
        else:
            msg = f"⚠️ Error al consultar la API de GitHub (Código HTTP {response.status_code})."
    except Exception as e:
        msg = f"❌ Excepción al consultar el estado: {e}"

    keyboard = [[InlineKeyboardButton("« Volver al Menú Principal", callback_data="menu_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if query:
        await query.message.edit_text(msg, reply_markup=reply_markup, parse_mode="Markdown", disable_web_page_preview=True)
    else:
        await update.message.reply_text(msg, reply_markup=reply_markup, parse_mode="Markdown", disable_web_page_preview=True)


def generar_excel_mes(noticias, year, month) -> io.BytesIO:
    libro = Workbook()
    hoja = libro.active
    hoja.title = f"{year}-{month}"

    encabezados = ["Fecha", "Categoría", "Titular", "Enlace"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    for item in noticias:
        hoja.append([item["fecha"], item["etiqueta"], item["titular"], item["link"]])
        fila = hoja.max_row
        celda_enlace = hoja.cell(row=fila, column=4)
        celda_enlace.hyperlink = item["link"]
        celda_enlace.font = Font(color="0563C1", underline="single")

    hoja.column_dimensions["A"].width = 14
    hoja.column_dimensions["B"].width = 16
    hoja.column_dimensions["C"].width = 70
    hoja.column_dimensions["D"].width = 55
    hoja.freeze_panes = "A2"

    buffer_bytes = io.BytesIO()
    libro.save(buffer_bytes)
    buffer_bytes.seek(0)
    buffer_bytes.name = f"dakhla_registro_{year}-{month}.xlsx"
    return buffer_bytes


async def manejar_botones(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "menu_main":
        await comando_menu(update, context)
        return
    elif data == "menu_ayuda":
        await comando_ayuda(update, context)
        return
    elif data == "menu_historico":
        await comando_registro(update, context)
        return
    elif data == "menu_exportar":
        await comando_exportar(update, context)
        return
    elif data == "menu_forzar":
        await comando_forzar(update, context)
        return
    elif data == "menu_estado":
        await comando_estado(update, context)
        return
    elif data == "menu_resumenmes":
        await comando_resumen_mensual(update, context)
        return

    datos = obtener_datos_registro()
    if not datos:
        await query.edit_message_text("⚠️ No se pudo cargar la información del registro.")
        return

    if data.startswith("year_"):
        year = data.split("_")[1]
        meses = datos.get(year, {})

        keyboard = []
        for month in sorted(meses.keys(), reverse=True):
            total_noticias = len(meses[month])
            clave_mes = str(month).zfill(2)
            nombre_mes = MESES_NOMBRE.get(clave_mes, f"Mes {month}")

            keyboard.append([
                InlineKeyboardButton(
                    f"📂 {nombre_mes} ({total_noticias} noticias)",
                    callback_data=f"month_{year}_{month}"
                )
            ])

        keyboard.append([InlineKeyboardButton("🔙 Volver a Años", callback_data="menu_historico")])
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            f"📅 *AÑO {year}*\n\nSelecciona un mes para ver el desglose:",
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

    elif data.startswith("month_"):
        _, year, month = data.split("_")
        noticias = datos.get(year, {}).get(month, [])

        clave_mes = str(month).zfill(2)
        nombre_mes_txt = MESES_NOMBRE.get(clave_mes, month).upper()

        if not noticias:
            keyboard = [[InlineKeyboardButton(f"🔙 Volver a Meses ({year})", callback_data=f"year_{year}")]]
            await query.edit_message_text(
                "No hay noticias registradas en este mes.",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return

        bloques = []
        texto_actual = f"📅 *REGISTRO — {nombre_mes_txt} {year}*\n\n"
        fecha_actual = ""

        for item in noticias:
            linea_fecha = ""
            if item["fecha"] != fecha_actual:
                fecha_actual = item["fecha"]
                linea_fecha = f"\n🗓️ *{fecha_actual}*\n"

            linea_noticia = f"{item['bandera']} {item['titular']} — [Noticia]({item['link']})\n"
            bloque_temp = linea_fecha + linea_noticia

            if len(texto_actual) + len(bloque_temp) > 3800:
                bloques.append(texto_actual)
                texto_actual = f"📅 *REGISTRO — {nombre_mes_txt} {year} (Cont.)*\n\n" + bloque_temp
            else:
                texto_actual += bloque_temp

        bloques.append(texto_actual)

        keyboard = [[InlineKeyboardButton(f"🔙 Volver a Meses ({year})", callback_data=f"year_{year}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)

        if len(bloques) > 1:
            for b in bloques[:-1]:
                await query.message.reply_text(b, parse_mode="Markdown", disable_web_page_preview=True)

            await query.message.reply_text(
                bloques[-1],
                reply_markup=reply_markup,
                parse_mode="Markdown",
                disable_web_page_preview=True
            )
        else:
            await query.edit_message_text(
                bloques[0],
                reply_markup=reply_markup,
                parse_mode="Markdown",
                disable_web_page_preview=True
            )

    elif data.startswith("exp_year_"):
        year = data.split("_")[2]
        meses = datos.get(year, {})

        keyboard = []
        for month in sorted(meses.keys(), reverse=True):
            total_noticias = len(meses[month])
            clave_mes = str(month).zfill(2)
            nombre_mes = MESES_NOMBRE.get(clave_mes, f"Mes {month}")

            keyboard.append([
                InlineKeyboardButton(
                    f"📤 {nombre_mes} ({total_noticias} noticias)",
                    callback_data=f"exp_month_{year}_{month}"
                )
            ])

        keyboard.append([InlineKeyboardButton("🔙 Volver a Años", callback_data="menu_exportar")])
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            f"📤 *Exportar — AÑO {year}*\n\nSelecciona el mes a exportar:",
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

    elif data.startswith("exp_month_"):
        _, _, year, month = data.split("_")
        noticias = datos.get(year, {}).get(month, [])

        if not noticias:
            await query.answer("No hay noticias registradas en ese mes.", show_alert=True)
            return

        archivo_excel = generar_excel_mes(noticias, year, month)
        clave_mes = str(month).zfill(2)
        nombre_mes = MESES_NOMBRE.get(clave_mes, month)

        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=InputFile(archivo_excel, filename=archivo_excel.name),
            caption=f"📤 Registro de {nombre_mes} {year} ({len(noticias)} noticias)"
        )
        await query.answer("Archivo enviado ✅")

    # --- Resumen mensual: Año -> Mes -> submenú (General / Valoración) ---
    elif data.startswith("resmes_year_"):
        year = data.split("_")[2]
        meses = datos.get(year, {})

        keyboard = []
        for month in sorted(meses.keys(), reverse=True):
            total_noticias = len(meses[month])
            clave_mes = str(month).zfill(2)
            nombre_mes = MESES_NOMBRE.get(clave_mes, f"Mes {month}")
            keyboard.append([
                InlineKeyboardButton(
                    f"📂 {nombre_mes} ({total_noticias} noticias)",
                    callback_data=f"resmes_month_{year}_{month}"
                )
            ])

        keyboard.append([InlineKeyboardButton("🔙 Volver a Años", callback_data="menu_resumenmes")])
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            f"⚓ *Resúmenes Mensuales — AÑO {year}*\n\nSelecciona el mes:",
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

    elif data.startswith("resmes_month_"):
        _, _, year, month = data.split("_")
        clave_mes = str(month).zfill(2)
        nombre_mes = MESES_NOMBRE.get(clave_mes, month)

        keyboard = [
            [InlineKeyboardButton("▶️ Resumen General del Mes", callback_data=f"resmes_general_{year}_{month}")],
            [InlineKeyboardButton("🕵️ Valoración de " + nombre_mes, callback_data=f"resmes_valoracion_{year}_{month}")],
            [InlineKeyboardButton("🔙 Volver a Meses", callback_data=f"resmes_year_{year}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        texto = (
            f"⚓ *Resúmenes Mensuales — {nombre_mes.upper()} {year}* ⚓\n\n"
            f"▶️ *Resumen General del Mes*\n"
            f"(Resumen de las noticias del mes sin valoración)\n\n"
            f"🕵️ *Valoración de {nombre_mes}*\n"
            f"(Resumen de las noticias del mes con enfoque de analista de inteligencia, sin valoraciones)"
        )
        await query.edit_message_text(texto, reply_markup=reply_markup, parse_mode="Markdown")

    elif data.startswith("resmes_general_") or data.startswith("resmes_valoracion_"):
        es_valoracion = data.startswith("resmes_valoracion_")
        prefijo = "resmes_valoracion_" if es_valoracion else "resmes_general_"
        year, month = data[len(prefijo):].split("_")
        items_mes = datos.get(year, {}).get(month, [])

        clave_mes = str(month).zfill(2)
        nombre_mes = MESES_NOMBRE.get(clave_mes, month)

        if not items_mes:
            await query.answer("No hay noticias registradas en ese mes.", show_alert=True)
            return

        await query.answer("Generando resumen, puede tardar unos segundos...")
        await query.message.reply_text(f"⏳ Generando el resumen de {nombre_mes} {year}, espera un momento...")

        texto_titulares = construir_texto_titulares_mes(items_mes)
        if es_valoracion:
            resumen = generar_resumen_mensual_analista(texto_titulares, nombre_mes, year)
            titulo = f"🕵️ <b>VALORACIÓN DE {html.escape(nombre_mes.upper())} {year}</b>"
        else:
            resumen = generar_resumen_mensual_general(texto_titulares, nombre_mes, year)
            titulo = f"▶️ <b>RESUMEN GENERAL DE {html.escape(nombre_mes.upper())} {year}</b>"

        if not resumen:
            await query.message.reply_text(
                "⚠️ No se ha podido generar el resumen (revisa que GROQ_API_KEY esté configurado en Render)."
            )
            return

        keyboard = [[InlineKeyboardButton("🔙 Volver", callback_data=f"resmes_month_{year}_{month}")]]
        mensaje_final = f"{titulo}\n\n{html.escape(resumen)}"
        for i in range(0, len(mensaje_final), 3900):
            trozo = mensaje_final[i:i + 3900]
            es_ultimo = i + 3900 >= len(mensaje_final)
            await query.message.reply_text(
                trozo,
                reply_markup=InlineKeyboardMarkup(keyboard) if es_ultimo else None,
                parse_mode="HTML"
            )


class _HealthCheckHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.end_headers()
        self.wfile.write("Bot de Telegram activo.".encode("utf-8"))

    def do_HEAD(self):
        self.do_GET()

    def log_message(self, format, *args):
        pass


def iniciar_servidor_salud():
    puerto = int(os.environ.get("PORT", "10000"))
    servidor = HTTPServer(("0.0.0.0", puerto), _HealthCheckHandler)
    logger.info(f"Servidor de salud escuchando en el puerto {puerto}")
    servidor.serve_forever()


def main():
    if not TELEGRAM_TOKEN:
        logger.error("No se ha configurado la variable de entorno TELEGRAM_TOKEN")
        return

    hilo_salud = threading.Thread(target=iniciar_servidor_salud, daemon=True)
    hilo_salud.start()

    hilo_reloj = threading.Thread(target=iniciar_reloj_disparo_diario, daemon=True)
    hilo_reloj.start()

    app = Application.builder().token(TELEGRAM_TOKEN).build()

    # Handlers para comandos y menú interactivo
    app.add_handler(CommandHandler("start", comando_menu))
    app.add_handler(CommandHandler("menu", comando_menu))
    app.add_handler(CommandHandler("ayuda", comando_ayuda))
    app.add_handler(CommandHandler("registro", comando_registro))
    app.add_handler(CommandHandler("historico", comando_registro))
    app.add_handler(CommandHandler("exportar", comando_exportar))
    app.add_handler(CommandHandler("resume_mes", comando_resumen_mensual))
    app.add_handler(CommandHandler("actualizar", comando_forzar))
    app.add_handler(CommandHandler("estado", comando_estado))
    app.add_handler(CallbackQueryHandler(manejar_botones))

    logger.info("Bot iniciado correctamente con panel completo y escuchando peticiones...")
    app.run_polling()


if __name__ == "__main__":
    main()
